from asyncio import wait_for

import page
from openpyxl.reader.excel import load_workbook
from playwright.sync_api import Page

class AfterSetup():

    def signup(self, page: Page, email: str, name: str, password: str, confirm_pass: str):
        page.locator("input[name='login']").fill(email)
        page.locator("input[name='name']").fill(name)
        page.locator("input[name='password']").fill(password)
        page.locator("input[name='confirm_password']").fill(confirm_pass)# Fixed variable name
        page.get_by_role("button", name="Sign Up").click()
        # try:
        #     error_locator = page.locator(".alert.alert-danger")
        #     if error_locator.is_visible():
        #         error_message = error_locator.text_content()
        #         assert False, f"Test Failed: {error_message}"
        #     else:
        #         print("Login Successfully")
        # except Exception as e:
        #     print(f"Error: {e}")
        #     page.close()



    def login_customer(self, page: Page):
        """Automates admin login using Playwright"""
        page.get_by_placeholder("Email").fill("portal")
        page.get_by_placeholder("Password").fill("portal")
        page.locator("button.btn.btn-primary").click()
        print(page.url)

    def login_seller(self, page: Page , email:str,password:str):
        """Automates admin login using Playwright"""
        page.get_by_placeholder("Email").fill(email)
        page.get_by_placeholder("Password").fill(password)
        page.locator("button.btn.btn-primary").click()
        print(page.url)


    def login_specific_seller(self, page: Page):
        """Automates admin login using Playwright"""
        page.get_by_placeholder("Email").fill("test@test.com")
        page.get_by_placeholder("Password").fill("webkul")
        page.locator("button.btn.btn-primary").click()
        print(page.url)

    def logout_seller(self,page:Page):
        logout = page.wait_for_selector("button.py-1.py-lg-0.o-dropdown.dropdown-toggle.dropdown", timeout=5000)
        logout.click()
        page.locator("//a[@data-menu='logout']").click()
        print("Successfully Logout the seller account")

    def logout_customer(self,page:Page):
        logout_dropdown = page.wait_for_selector(
            "ul.navbar-nav.align-items-center.gap-2.flex-shrink-0.justify-content-end.ps-3 li.dropdown.o_no_autohide_item",
            timeout=5000
        )
        logout_dropdown.click()
        logout_btn = page.locator("div.dropdown-menu.js_usermenu.dropdown-menu-end.show a#o_logout")
        logout_btn.click()
        print("✅ Successfully logged out the customer account")


    def login_admin(self, page: Page):
        """Automates admin login using Playwright"""
        page.get_by_placeholder("Email").fill("admin")
        page.get_by_placeholder("Password").fill("webkul")
        page.locator("button.btn.btn-primary").click()
        print(page.url)


def get_seller_credentials(Name):
    wb = load_workbook('User.xlsx')
    sheet = wb.active
    print("Reading credentials from Excel...")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == Name:
            return {"email": row[1], "password": row[2]}
    return None




